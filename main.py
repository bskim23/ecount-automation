# main.py
# APP_REV: 2026-02-24_15_full_debug_v2

APP_REV = "2026-02-24_15_full_debug_v2"

from flask import Flask, request, jsonify
import os, json, base64, re, datetime, io, traceback
from typing import Dict, Any, Tuple, List, Optional

import gspread
from google.oauth2.service_account import Credentials

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError
    PLAYWRIGHT_IMPORT_OK = True
except Exception:
    PLAYWRIGHT_IMPORT_OK = False

app = Flask(__name__)

# =============================================================================
# 공통 유틸
# =============================================================================

def now_kst_str() -> str:
    kst = datetime.timezone(datetime.timedelta(hours=9))
    return datetime.datetime.now(tz=kst).strftime("%Y-%m-%d %H:%M:%S%z")

def get_env(name: str, default: str = "") -> str:
    v = os.environ.get(name, default)
    return v if v is not None else default

def safe_str(x: Any, limit: int = 2000) -> str:
    s = ""
    try:
        s = str(x)
    except Exception:
        s = repr(x)
    if len(s) > limit:
        return s[:limit] + "...(truncated)"
    return s

def b64_png(page) -> str:
    try:
        return base64.b64encode(page.screenshot(full_page=True)).decode("utf-8")
    except Exception:
        return ""

def frames_snapshot(page, selector_map: Dict[str, str]) -> Dict[str, Any]:
    """현재 page.frames 전체에 대해 url/readyState/selector count를 덤프"""
    out = {"page_url": page.url, "page_hash": "", "frames": []}
    try:
        out["page_hash"] = page.evaluate("location.hash")
    except Exception:
        out["page_hash"] = ""
    frames = page.frames
    for i, fr in enumerate(frames):
        item = {"i": i, "name": fr.name, "url": fr.url, "readyState": "", "selector_counts": {}}
        try:
            item["readyState"] = fr.evaluate("document.readyState")
        except Exception:
            item["readyState"] = "?"
        for k, sel in selector_map.items():
            try:
                item["selector_counts"][k] = fr.locator(sel).count()
            except Exception:
                item["selector_counts"][k] = -1
        out["frames"].append(item)
    return out

def attach_debug(result: Dict[str, Any], page, stage: str, extra: Optional[Dict[str, Any]] = None, selector_map: Optional[Dict[str, str]] = None):
    if "debug" not in result:
        result["debug"] = {}
    if selector_map is None:
        selector_map = {}
    snap = {
        "stage": stage,
        "timestamp": now_kst_str(),
        "page_url": getattr(page, "url", ""),
        "frames_state": frames_snapshot(page, selector_map) if page else {},
        "screenshot_b64": b64_png(page) if page else "",
    }
    if extra:
        snap["extra"] = extra
    result["debug"][stage] = snap

def env_health() -> Dict[str, Any]:
    keys = [
        "GOOGLE_SERVICE_ACCOUNT_JSON", "GOOGLE_SHEET_ID", "SHEET_NAME",
        "ECOUNT_LOGIN_URL", "ECOUNT_ERP_URL",
        "COM_CODE", "USER_ID", "USER_PW",
    ]
    out = {"app_rev": APP_REV, "timestamp": now_kst_str(), "ok": True, "env": {}}
    for k in keys:
        v = get_env(k, "")
        out["env"][k] = ("✅" if (v and len(v.strip()) > 0) else "❌")
    if not PLAYWRIGHT_IMPORT_OK:
        out["ok"] = False
        out["playwright"] = "import_failed"
    else:
        out["playwright"] = "ok"
    return out

# =============================================================================
# 구글 시트
# =============================================================================

def gspread_client() -> gspread.Client:
    raw = get_env("GOOGLE_SERVICE_ACCOUNT_JSON").strip()
    # JSON 또는 base64(JSON) 대응
    try:
        info = json.loads(raw) if raw.startswith("{") else json.loads(base64.b64decode(raw).decode("utf-8"))
    except Exception:
        info = json.loads(raw)
    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    return gspread.authorize(Credentials.from_service_account_info(info, scopes=scopes))

def open_target_worksheet(gc: gspread.Client):
    sh = gc.open_by_key(get_env("GOOGLE_SHEET_ID").strip())
    ws = sh.worksheet(get_env("SHEET_NAME", "SAT Raw").strip())
    return sh, ws

# =============================================================================
# 날짜 / 월키
# =============================================================================

def ym_key_from_a(a_val: Any) -> str:
    m = re.search(r"(\d{4})[/-](\d{2})[/-](\d{2})", str(a_val or ""))
    return f"{m.group(1)}/{m.group(2)}" if m else ""

# =============================================================================
# Playwright: 클릭/탐색 유틸 (iframe 다중 스캔)
# =============================================================================

def find_and_click_css_in_frames(page, css: str, timeout_ms: int = 30000, force: bool = True) -> Tuple[bool, str]:
    """모든 프레임에서 css를 찾아 보이는 첫 요소 클릭"""
    deadline = datetime.datetime.now() + datetime.timedelta(milliseconds=timeout_ms)
    last_err = None
    while datetime.datetime.now() < deadline:
        for fr in page.frames:
            try:
                loc = fr.locator(css)
                if loc.count() > 0:
                    el = loc.first
                    if el.is_visible():
                        el.click(force=force)
                        return True, f"CLICKED:{css}"
            except Exception as e:
                last_err = safe_str(e)
                continue
        page.wait_for_timeout(350)
    return False, f"TIMEOUT:{css} last_err={last_err}"

def find_and_click_text_in_frames(page, text_list: List[str], exact: bool = True, timeout_ms: int = 30000, force: bool = True) -> Tuple[bool, str]:
    """모든 프레임에서 텍스트 기반 클릭(여러 후보 지원)"""
    deadline = datetime.datetime.now() + datetime.timedelta(milliseconds=timeout_ms)
    last_err = None
    while datetime.datetime.now() < deadline:
        for fr in page.frames:
            for t in text_list:
                try:
                    loc = fr.get_by_text(t, exact=exact)
                    if loc.count() > 0:
                        el = loc.first
                        if el.is_visible():
                            el.click(force=force)
                            return True, f"CLICKED_TEXT:{t}"
                except Exception as e:
                    last_err = safe_str(e)
                    continue
        page.wait_for_timeout(350)
    return False, f"TIMEOUT_TEXT:{text_list} last_err={last_err}"

def collect_login_error_text(page) -> str:
    """로그인 페이지/상태에서 실패 메시지 후보를 최대한 긁어오기"""
    patterns = [
        r"비밀번호", r"아이디", r"오류", r"실패", r"인증", r"확인", r"차단", r"보안", r"다시", r"틀렸",
        r"Password", r"ID", r"Error", r"Failed", r"Verify", r"Blocked", r"Security"
    ]
    combined = "|".join(patterns)
    txt = ""
    try:
        # 화면에 뜨는 텍스트 중 패턴 매칭되는 줄만 수집
        hits = page.locator(f"text=/{combined}/i").all_inner_texts()
        hits = [h.strip() for h in hits if h and h.strip()]
        if hits:
            txt = " | ".join(hits[:10])
    except Exception:
        pass
    # 흔한 에러 박스 후보
    for sel in [".error", ".msg", "#msg", ".login_error", ".warning", ".notice", ".alert", ".toast"]:
        try:
            if page.locator(sel).count() > 0:
                t = page.locator(sel).first.inner_text().strip()
                if t:
                    txt = (txt + " | " + t).strip(" |")
        except Exception:
            pass
    return safe_str(txt, 2000)

# =============================================================================
# ECOUNT: ERP 다운로드/파싱
# =============================================================================

def ecount_download_and_parse() -> Tuple[bool, Dict[str, Any]]:
    if not PLAYWRIGHT_IMPORT_OK:
        return False, {"error": "Playwright import failed", "app_rev": APP_REV}

    result: Dict[str, Any] = {"app_rev": APP_REV, "steps": {}, "timestamp": now_kst_str()}
    selector_map = {
        "login_com": "#com_code",
        "login_id": "#id",
        "login_pw": "#passwd",
        "inv": "a#link_depth1_MENUTREE_000004",
        "sales_mgmt": "a[href*='menuSeq=MENUTREE_000030'][href*='depth=2']",
        "sales_status": "a#link_depth4_MENUTREE_000494",
        "sat": "a.btn.btn-primary.btn-fn[data-own-layer-box-id]",
        "range_span": "span:has-text('금월(~오늘)')",
        "excel_btn": "span:has-text('Excel(화면)')",
    }

    login_url = get_env("ECOUNT_LOGIN_URL", "https://login.ecount.com/Login/").strip()
    # IMPORTANT: ec_req_sid 같은 값 절대 하드코딩 금지. base ERP URL만.
    erp_url = get_env("ECOUNT_ERP_URL", "https://loginca.ecount.com/ec5/view/erp?w_flag=1").strip()

    com_code = get_env("COM_CODE").strip()
    user_id = get_env("USER_ID").strip()
    user_pw = get_env("USER_PW").strip()

    if not (com_code and user_id and user_pw):
        return False, {"error": "Missing COM_CODE/USER_ID/USER_PW", "app_rev": APP_REV}

    try:
        from openpyxl import load_workbook

        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage"]
            )
            context = browser.new_context(
                accept_downloads=True,
                viewport={"width": 1920, "height": 1080},
                # 헤드리스 탐지 완화 목적(완전 해결책은 아니지만 도움이 되는 경우가 있음)
                user_agent=(
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/122.0.0.0 Safari/537.36"
                )
            )
            page = context.new_page()
            page.set_default_timeout(60000)

            # 콘솔/페이지 에러 수집
            console_logs: List[Dict[str, Any]] = []
            page_errors: List[str] = []
            req_failed: List[Dict[str, Any]] = []

            def on_console(msg):
                try:
                    console_logs.append({"type": msg.type, "text": safe_str(msg.text, 400)})
                except Exception:
                    pass

            def on_page_error(err):
                try:
                    page_errors.append(safe_str(err, 800))
                except Exception:
                    pass

            def on_req_failed(req):
                try:
                    req_failed.append({
                        "url": safe_str(req.url, 300),
                        "method": req.method,
                        "failure": safe_str(req.failure, 300)
                    })
                except Exception:
                    pass

            page.on("console", on_console)
            page.on("pageerror", on_page_error)
            page.on("requestfailed", on_req_failed)

            # 0) ERP URL 먼저 열기 → 세션 없으면 login으로 튕기도록(정상 SSO 흐름 유도)
            page.goto(erp_url, wait_until="load")
            page.wait_for_timeout(800)

            # 1) login 페이지로 왔는지 확인 → 아니면 login_url로 강제 이동
            if "login.ecount.com" not in page.url:
                page.goto(login_url, wait_until="load")
                page.wait_for_timeout(500)

            attach_debug(result, page, "login_page_loaded", selector_map=selector_map)

            # 2) 로그인 폼 채우기
            page.locator("#com_code").fill(com_code)
            page.locator("#id").fill(user_id)
            page.locator("#passwd").fill(user_pw)

            # 3) Enter 제출은 실패가 잦아서 "submit 버튼 클릭 우선" + fallback Enter
            login_clicked = False
            # 사이트별로 버튼 셀렉터가 달라질 수 있어 후보를 여러 개 둠
            login_btn_selectors = [
                "button[type='submit']",
                "input[type='submit']",
                "button:has-text('로그인')",
                "button:has-text('Login')",
                "#btn_login",
                ".btn_login",
            ]

            for sel in login_btn_selectors:
                try:
                    if page.locator(sel).count() > 0:
                        with page.expect_navigation(timeout=60000):
                            page.locator(sel).first.click()
                        login_clicked = True
                        break
                except Exception:
                    continue

            if not login_clicked:
                # 최후 수단: Enter
                try:
                    with page.expect_navigation(timeout=60000):
                        page.keyboard.press("Enter")
                except Exception:
                    page.keyboard.press("Enter")

            # 4) 로그인 후 "ERP 도메인(loginca)"으로 넘어갔는지 확정 판정
            try:
                page.wait_for_url("**loginca.ecount.com/**", timeout=60000)
                result["steps"]["login"] = "done"
            except Exception:
                # 아직 login이면: 에러 텍스트/쿠키/로그 기록 + 캡처
                err_txt = collect_login_error_text(page)
                cookies = []
                try:
                    cookies = context.cookies()
                except Exception:
                    cookies = []
                attach_debug(
                    result,
                    page,
                    "login_not_redirected",
                    extra={
                        "error_text": err_txt,
                        "cookie_count": len(cookies),
                        "cookies_sample": cookies[:5],
                        "console_tail": console_logs[-15:],
                        "page_errors": page_errors[-10:],
                        "req_failed_tail": req_failed[-10:],
                    },
                    selector_map=selector_map
                )
                browser.close()
                return False, {"error": "Login did not redirect to ERP domain", "partial": result}

            # 5) ERP base 진입 재확인(세션이 생겼다면 여기서 ERP가 열려야 함)
            page.goto(erp_url, wait_until="load")
            page.wait_for_timeout(1000)
            try:
                page.wait_for_url("**/ec5/view/erp**", timeout=60000)
            except Exception:
                err_txt = collect_login_error_text(page)
                attach_debug(
                    result, page, "erp_entry_failed",
                    extra={"error_text": err_txt, "console_tail": console_logs[-15:], "req_failed_tail": req_failed[-10:]},
                    selector_map=selector_map
                )
                browser.close()
                return False, {"error": "ERP entry failed after login", "partial": result}

            attach_debug(result, page, "erp_loaded", selector_map=selector_map)

            # 6) 메뉴 클릭 순서 확정 (재고 I → 영업관리 → 판매현황)
            ok, msg = find_and_click_css_in_frames(page, "a#link_depth1_MENUTREE_000004", timeout_ms=40000)
            result["steps"]["click_inv"] = msg
            if not ok:
                attach_debug(result, page, "fail_click_inv", extra={"css": "a#link_depth1_MENUTREE_000004"}, selector_map=selector_map)
                browser.close()
                return False, {"error": f"Timeout click: 재고 I(a#link_depth1_MENUTREE_000004)", "partial": result}

            page.wait_for_timeout(800)

            # 영업관리: 동일 항목이 2개 보일 수 있다고 하셨으므로
            # 1) id 기반 후보
            # 2) href 기반 후보
            # 3) 텍스트 기반 후보
            sales_mgmt_css_candidates = [
                "a#link_depth2_MENUTREE_000030",
                "a[href*='menuSeq=MENUTREE_000030'][href*='depth=2']",
            ]
            clicked = False
            last = ""
            for css in sales_mgmt_css_candidates:
                ok, msg = find_and_click_css_in_frames(page, css, timeout_ms=25000)
                last = msg
                if ok:
                    clicked = True
                    break
            if not clicked:
                ok, msg = find_and_click_text_in_frames(page, ["영업관리"], exact=True, timeout_ms=25000)
                last = msg
                clicked = ok
            result["steps"]["click_sales_mgmt"] = last
            if not clicked:
                attach_debug(result, page, "fail_click_sales_mgmt", extra={"candidates": sales_mgmt_css_candidates}, selector_map=selector_map)
                browser.close()
                return False, {"error": "Timeout click: 영업관리", "partial": result}

            page.wait_for_timeout(800)

            # 판매현황: 사용자 확인 anchor id
            ok, msg = find_and_click_css_in_frames(page, "a#link_depth4_MENUTREE_000494", timeout_ms=40000)
            result["steps"]["click_sales_status"] = msg
            if not ok:
                attach_debug(result, page, "fail_click_sales_status", extra={"css": "a#link_depth4_MENUTREE_000494"}, selector_map=selector_map)
                browser.close()
                return False, {"error": "Timeout click: 판매현황(a#link_depth4_MENUTREE_000494)", "partial": result}

            page.wait_for_timeout(1200)

            # 7) SAT 버튼: <a class="btn btn-primary btn-fn" data-own-layer-box-id="layer_5_50863"></a>
            # data-own-layer-box-id는 환경에 따라 달라질 수 있어
            # 1) 정확 id가 있으면 우선
            # 2) 없으면 btn-fn 중 "가시" 요소를 클릭 (위험하니 debug를 많이 남김)
            sat_css_candidates = [
                "a.btn.btn-primary.btn-fn[data-own-layer-box-id='layer_5_50863']",
                "a.btn.btn-primary.btn-fn[data-own-layer-box-id]",
            ]
            clicked = False
            last = ""
            for css in sat_css_candidates:
                ok, msg = find_and_click_css_in_frames(page, css, timeout_ms=30000)
                last = msg
                if ok:
                    clicked = True
                    break
            result["steps"]["click_sat"] = last
            if not clicked:
                attach_debug(result, page, "fail_click_sat", extra={"candidates": sat_css_candidates}, selector_map=selector_map)
                browser.close()
                return False, {"error": "Timeout click: SAT 버튼", "partial": result}

            page.wait_for_timeout(1500)

            # 8) 기간: 금월(~오늘) span 1개라고 하셨음
            ok, msg = find_and_click_text_in_frames(page, ["금월(~오늘)"], exact=True, timeout_ms=30000)
            result["steps"]["click_range"] = msg
            if not ok:
                # 텍스트 클릭 실패하면 span 셀렉터로 재시도
                ok2, msg2 = find_and_click_css_in_frames(page, "span:has-text('금월(~오늘)')", timeout_ms=20000)
                result["steps"]["click_range_css"] = msg2
                ok = ok2
            if not ok:
                attach_debug(result, page, "fail_click_range", extra={"text": "금월(~오늘)"}, selector_map=selector_map)
                browser.close()
                return False, {"error": "Timeout click: 금월(~오늘)", "partial": result}

            page.wait_for_timeout(1800)

            # 9) Excel(화면) 다운로드: 같은 html이 2개라 "가시 요소" 우선
            # expect_download로 실제 다운로드가 트리거되는지 확인
            def click_excel():
                # 1) span 텍스트 기반
                ok, msg = find_and_click_text_in_frames(page, ["Excel(화면)", "엑셀(화면)", "Excel"], exact=True, timeout_ms=25000)
                result["steps"]["click_excel_text"] = msg
                if ok:
                    return True
                # 2) css 기반
                ok, msg = find_and_click_css_in_frames(page, "span:has-text('Excel(화면)')", timeout_ms=15000)
                result["steps"]["click_excel_css"] = msg
                return ok

            try:
                with page.expect_download(timeout=90000) as dlinfo:
                    ok = click_excel()
                    if not ok:
                        raise RuntimeError("Excel button not found/clickable")
                download = dlinfo.value

                # 10) 파일을 메모리로 읽기
                file_buffer = io.BytesIO()
                with download.create_read_stream() as stream:
                    while True:
                        chunk = stream.read(1024 * 128)
                        if not chunk:
                            break
                        file_buffer.write(chunk)
                file_buffer.seek(0)

            except Exception as e:
                attach_debug(
                    result, page, "excel_download_failed",
                    extra={"err": safe_str(e), "console_tail": console_logs[-15:], "req_failed_tail": req_failed[-10:]},
                    selector_map=selector_map
                )
                browser.close()
                return False, {"error": f"Download Fail: {safe_str(e)}", "partial": result}

            # 11) 엑셀 파싱
            wb = load_workbook(file_buffer, data_only=True, read_only=True)
            ws = wb.active

            rows: List[List[Any]] = []
            # 헤더 2줄 제외, 하단 합계 3줄 제외 (기존 로직 유지)
            for r in range(3, ws.max_row - 2):
                row_val = [ws.cell(row=r, column=c).value for c in range(1, 11)]
                if row_val[0]:
                    rows.append(row_val)

            result["row_count"] = len(rows)
            result["month_key"] = datetime.datetime.now().strftime("%Y/%m")
            if rows:
                m = re.search(r"(\d{4})[/-](\d{2})[/-](\d{2})", str(rows[0][0]))
                if m:
                    result["month_key"] = f"{m.group(1)}/{m.group(2)}"
            result["rows"] = rows

            # 로그 tail 보관
            result["debug_tail"] = {
                "console_tail": console_logs[-30:],
                "page_errors": page_errors[-10:],
                "req_failed_tail": req_failed[-20:],
            }

            browser.close()
            return True, result

    except Exception as e:
        # 예외가 났을 때도 partial result를 함께 반환
        tb = traceback.format_exc()
        return False, {"error": f"Execution Error: {safe_str(e)}", "traceback": safe_str(tb, 4000), "partial": result}

# =============================================================================
# Stage: gsheet 업데이트
# =============================================================================

def stage_gsheet_update(new_rows: List[List[Any]], month_key: str) -> Dict[str, Any]:
    gc = gspread_client()
    sh, ws = open_target_worksheet(gc)

    all_vals = ws.get_all_values()
    header = all_vals[0] if all_vals else ["일자-No.", "품목명(규격)", "수량", "단가", "공급가액", "부가세", "합계", "거래처명", "적요", "거래처계층그룹명"]
    body = all_vals[1:]

    kept = [r for r in body if ym_key_from_a(r[0]) != month_key]

    ws.clear()
    ws.update("A1", [header] + kept + new_rows, value_input_option="USER_ENTERED")

    return {"ok": True, "month": month_key, "count": len(new_rows), "timestamp": now_kst_str()}

# =============================================================================
# Stage 라우팅
# =============================================================================

def stage_env() -> Dict[str, Any]:
    return env_health()

def stage_erp() -> Dict[str, Any]:
    ok, erp_res = ecount_download_and_parse()
    if not ok:
        return {"ok": False, "error": erp_res}
    return {"ok": True, "month": erp_res.get("month_key"), "count": erp_res.get("row_count"), "timestamp": now_kst_str(), "erp": {"month_key": erp_res.get("month_key"), "row_count": erp_res.get("row_count")}}

def stage_all() -> Dict[str, Any]:
    ok, erp_res = ecount_download_and_parse()
    if not ok:
        return {"ok": False, "error": erp_res}

    month_key = erp_res["month_key"]
    new_rows = erp_res["rows"]

    try:
        gs = stage_gsheet_update(new_rows, month_key)
        gs["erp_debug"] = erp_res.get("debug", {})
        return gs
    except Exception as e:
        return {"ok": False, "error": f"GSHEET Error: {safe_str(e)}", "erp_debug": erp_res.get("debug", {})}

# =============================================================================
# Flask endpoints
# =============================================================================

@app.route("/")
def health():
    return f"OK | {APP_REV} | {now_kst_str()}", 200

@app.route("/run", methods=["GET"])
def run_job():
    stage = (request.args.get("stage") or "all").strip().lower()

    if stage in ("", "help"):
        return jsonify({
            "ok": True,
            "app_rev": APP_REV,
            "timestamp": now_kst_str(),
            "stages": ["env", "erp", "gsheet", "all"],
            "examples": [
                "/run?stage=env",
                "/run?stage=erp",
                "/run?stage=all",
            ],
        }), 200

    if stage == "env":
        res = stage_env()
        return jsonify(res), (200 if res.get("ok") else 500)

    if stage == "erp":
        # erp만 실행해서 디버그를 빠르게 보는 용도
        ok, erp_res = ecount_download_and_parse()
        return jsonify({"ok": ok, "error": (None if ok else erp_res), "erp": (erp_res if ok else None)}), (200 if ok else 500)

    if stage == "gsheet":
        # gsheet만 단독으로는 의미가 약해서 막아둠(필요하면 구현)
        return jsonify({"ok": False, "error": "gsheet stage requires ERP rows. Use stage=all."}), 400

    # default: all
    res = stage_all()
    return jsonify(res), (200 if res.get("ok") else 500)

if __name__ == "__main__":
    # Cloud Run에서는 보통 gunicorn이 실행하지만 로컬 테스트를 위해 유지
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8080)), threaded=True)
